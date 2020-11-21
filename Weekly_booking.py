from openpyxl import load_workbook

weekly_report_template = "OPERA_Weekly_Hours_19_20_v10_FWxx_SW_VV.xlsx"

month = input("Enter the Month : ") 
week_num = input("Enter the Week Number : ")
first_day = input("Enter the First day of the current week : ")

weekly_report = "OPERA_Weekly_Hours_19_20_v10_FW" +week_num+ "_SW_VV.xlsx"

#Open workbook
workbook = load_workbook(weekly_report_template)
workbook.sheetnames
sheet = workbook.active
sheet.title = "Wk_" + month +"_"+ first_day

#Clear existing data in cell
for row in sheet["E5:Z22"]:
    for cell in row:
        cell.value = None

class SW:
    def sai():
        Cyber_SW = input("Enter the hours of Sai booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of Sai booked in SW_Sustain: ")
        Leave = input("Enter the hours of Sai booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")

        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E5"] = float(Cyber_SW)
            sheet["Q5"] = float(SW_Sustain)
            sheet["V5"] = float(Leave)

        if(Alista == "y"):
            sheet["AC5"] = "Yes"
        else:
            sheet["AC5"] = "No"

        print("The total hours booked by Sai for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))

    def haripriya():
        Cyber_SW = input("Enter the hours of Haripriya booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of Haripriya booked in SW_Sustain: ")
        Leave = input("Enter the hours of Haripriya booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")

        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E6"] = float(Cyber_SW)
            sheet["Q6"] = float(SW_Sustain)
            sheet["V6"] = float(Leave)

        if(Alista == "y"):
            sheet["AC6"] = "Yes"
        else:
            sheet["AC6"] = "No"
            
        print("The total hours booked by haripriya for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))

    def neeladri():
        Cyber_SW = input("Enter the hours of neeladri booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of neeladri booked in SW_Sustain: ")
        Leave = input("Enter the hours of neeladri booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")

        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E7"] = float(Cyber_SW)
            sheet["Q7"] = float(SW_Sustain)
            sheet["V7"] = float(Leave)

        if(Alista == "y"):
            sheet["AC7"] = "Yes"
        else:
            sheet["AC7"] = "No"
            
        print("The total hours booked by neeladri for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))

    def pankaj():
        Cyber_SW = input("Enter the hours of pankaj booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of pankaj booked in SW_Sustain: ")
        Leave = input("Enter the hours of pankaj booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")

        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E8"] = float(Cyber_SW)
            sheet["Q8"] = float(SW_Sustain)
            sheet["V8"] = float(Leave)

        if(Alista == "y"):
            sheet["AC8"] = "Yes"
        else:
            sheet["AC8"] = "No"
            
        print("The total hours booked by pankaj for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))

    def chandrappa():
        Cyber_SW = input("Enter the hours of chandrappa booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of chandrappa booked in SW_Sustain: ")
        Leave = input("Enter the hours of chandrappa booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")

        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E9"] = float(Cyber_SW)
            sheet["Q9"] = float(SW_Sustain)
            sheet["V9"] = float(Leave)

        if(Alista == "y"):
            sheet["AC9"] = "Yes"
        else:
            sheet["AC9"] = "No"
            
        print("The total hours booked by chandrappa for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))

    def anoopFrancis():
        Cyber_SW = input("Enter the hours of anoopFrancis booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of anoopFrancis booked in SW_Sustain: ")
        Leave = input("Enter the hours of anoopFrancis booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")
        
        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E10"] = float(Cyber_SW)
            sheet["Q10"] = float(SW_Sustain)
            sheet["V10"] = float(Leave)

        if(Alista == "y"):
            sheet["AC10"] = "Yes"
        else:
            sheet["AC10"] = "No"
            
        print("The total hours booked by anoopFrancis for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))

    def sowmya():
        Cyber_SW = input("Enter the hours of sowmya booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of sowmya booked in SW_Sustain: ")
        Leave = input("Enter the hours of sowmya booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")

        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E11"] = float(Cyber_SW)
            sheet["Q11"] = float(SW_Sustain)
            sheet["V11"] = float(Leave)

        if(Alista == "y"):
            sheet["AC11"] = "Yes"
        else:
            sheet["AC11"] = "No"
            
        print("The total hours booked by sowmya for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))

    def aravinda():
        Cyber_SW = input("Enter the hours of aravinda booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of aravinda booked in SW_Sustain: ")
        Leave = input("Enter the hours of aravinda booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")

        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E12"] = float(Cyber_SW)
            sheet["Q12"] = float(SW_Sustain)
            sheet["V12"] = float(Leave)

        if(Alista == "y"):
            sheet["AC12"] = "Yes"
        else:
            sheet["AC12"] = "No"
            
        print("The total hours booked by aravinda for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))

    def sw1():
        Cyber_SW = input("Enter the hours of Daniel booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of Daniel booked in SW_Sustain: ")
        Leave = input("Enter the hours of Daniel booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")

        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E13"] = float(Cyber_SW)
            sheet["Q13"] = float(SW_Sustain)
            sheet["V13"] = float(Leave)

        if(Alista == "y"):
            sheet["AC13"] = "Yes"
        else:
            sheet["AC13"] = "No"
            
        print("The total hours booked by Daniel for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))
	
    def sw2():
        Cyber_SW = input("Enter the hours of Mamatha booked in Cyber_SW: ")
        SW_Sustain = input("Enter the hours of Mamatha booked in SW_Sustain: ")
        Leave = input("Enter the hours of Mamatha booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")
        
        if (float(Cyber_SW)  or float(SW_Sustain) or float(Leave)) > 0:
            sheet["E14"] = float(Cyber_SW)
            sheet["Q14"] = float(SW_Sustain)
            sheet["V14"] = float(Leave)

        if(Alista == "y"):
            sheet["AC14"] = "Yes"
        else:
            sheet["AC14"] = "No"
            
        print("The total hours booked by Mamatha for the current week is:", float(Cyber_SW) + float(SW_Sustain) + float(Leave))
		
class VV:
    def jagan():
        Cyber_VV = input("Enter the hours of jagan booked in Cyber_VV: ")
        VV_Sustain = input("Enter the hours of jagan booked in VV_Sustain: ")
        Leave = input("Enter the hours of jagan booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")
        
        if (float(Cyber_VV)  or float(VV_Sustain) or float(Leave)) > 0:
            sheet["F15"] = float(Cyber_VV)
            sheet["R15"] = float(VV_Sustain)
            sheet["V15"] = float(Leave)

        if(Alista == "y"):
            sheet["AC15"] = "Yes"
        else:
            sheet["AC15"] = "No"
            
        print("The total hours booked by jagan for the current week is:", float(Cyber_VV) + float(VV_Sustain) + float(Leave))

    def vikas():
        Cyber_VV = input("Enter the hours of vikas booked in Cyber_VV: ")
        VV_Sustain = input("Enter the hours of vikas booked in VV_Sustain: ")
        Leave = input("Enter the hours of vikas booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")
        
        if (float(Cyber_VV)  or float(VV_Sustain) or float(Leave)) > 0:
            sheet["F16"] = float(Cyber_VV)
            sheet["R16"] = float(VV_Sustain)
            sheet["V16"] = float(Leave)

        if(Alista == "y"):
            sheet["AC16"] = "Yes"
        else:
            sheet["AC16"] = "No"
            
        print("The total hours booked by vikas for the current week is:", float(Cyber_VV) + float(VV_Sustain) + float(Leave))

    def bobba():
        Cyber_VV = input("Enter the hours of bobba booked in Cyber_VV: ")
        VV_Sustain = input("Enter the hours of bobba booked in VV_Sustain: ")
        Leave = input("Enter the hours of bobba booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")
        
        if (float(Cyber_VV)  or float(VV_Sustain) or float(Leave)) > 0:
            sheet["F17"] = float(Cyber_VV)
            sheet["R17"] = float(VV_Sustain)
            sheet["V17"] = float(Leave)

        if(Alista == "y"):
            sheet["AC17"] = "Yes"
        else:
            sheet["AC17"] = "No"
            
        print("The total hours booked by bobba for the current week is:", float(Cyber_VV) + float(VV_Sustain) + float(Leave))
    
    def anoopJose():
        Cyber_VV = input("Enter the hours of anoopJose booked in Cyber_VV: ")
        VV_Sustain = input("Enter the hours of anoopJose booked in VV_Sustain: ")
        Leave = input("Enter the hours of anoopJose booked in Leave: ")
        Alista = input("If Alista screenshot available: y/n ")
        
        if (float(Cyber_VV)  or float(VV_Sustain) or float(Leave)) > 0:
            sheet["F19"] = float(Cyber_VV)
            sheet["R19"] = float(VV_Sustain)
            sheet["V19"] = float(Leave)

        if(Alista == "y"):
            sheet["AC19"] = "Yes"
        else:
            sheet["AC19"] = "No"
            
        print("The total hours booked by anoopJose for the current week is:", float(Cyber_VV) + float(VV_Sustain) + float(Leave))

    def vv2():
        Cyber_VV = input("Enter the hours of poulamee booked in Cyber_VV: ")
        Leave = input("Enter the hours of poulamee booked in Leave: ")
        Training  = input("Enter the hours of poulamee booked in Training: ")
        Alista = input("If Alista screenshot available: y/n ")
        
        if (float(Cyber_VV) or float(Leave) or float(Training)) > 0:
            sheet["F20"] = float(Cyber_VV)
            sheet["Z20"] = float(Training)
            sheet["V20"] = float(Leave)
            
        if(Alista == "y"):
            sheet["AC20"] = "Yes"
        else:
            sheet["AC20"] = "No"
            
        print("The total hours booked by poulamee for the current week is:", float(Cyber_VV) + float(Leave) + float(Training))

    def vr_activities():
        Cyber_VV = input("Enter the hours of vr_activities booked in Cyber_VV: ")

        if (float(Cyber_VV)) > 0:
            sheet["F21"] = float(Cyber_VV)

        print("The total hours booked by vr_activities for the current week is:", float(Cyber_VV))
        
class safety:
    def swami():
        Cyber_Safety = input("Enter the hours of swami booked in Cyber_Safety: ")
        Alista = input("If Alista screenshot available: y/n ")
        
        if (float(Cyber_Safety)) > 0:
            sheet["G22"] = float(Cyber_Safety)
            
        if(Alista == "y"):
            sheet["AC22"] = "Yes"
        else:
            sheet["AC22"] = "No"
            
        print("The total hours booked by swami for the current week is:", float(Cyber_Safety))

class opera:
    def all_users():
        #For V&V Team
        VV.vr_activities()
        VV.anoopJose()
        VV.bobba()
        VV.jagan()
        VV.vv2()
        VV.vikas()
        #For Software Team
        SW.anoopFrancis()
        SW.sai()
        SW.neeladri()
        SW.chandrappa()
        SW.sowmya()
        SW.haripriya()
        SW.sw1()
        SW.sw2()
        #SW.aravinda()
        SW.pankaj()
        #For Safety Team
        safety.swami()


a = 0
while a<19:
    user_input = input('''
    Press 1 for vr_activities
    Press 2 for anoopJose
    Press 3 for bobba
    Press 4 for jagan
    Press 5 for poulamee
    Press 6 for vikas
    Press 7 for anoopFrancis
    Press 8 for Sai
    Press 9 for neeladri
    Press 10 for chandrappa
    Press 11 for sowmya
    Press 12 for haripriya
    Press 13 for daniel
    Press 14 for mamatha
    Press 15 for aravinda
    Press 16 for pankaj
    Press 17 for swami
    Press 18 to input full details
    Press 19 to exit
    Enter your choice: 
    ''')

    print("Entered user input is :", user_input)

    if(int(user_input) == 1):
        VV.vr_activities()
        a = int(user_input)

    elif(int(user_input) == 2):
        VV.anoopJose()
        a = int(user_input)

    elif(int(user_input) == 3):
        VV.bobba()
        a = int(user_input)
        
    elif(int(user_input) == 4):
        VV.jagan()
        a = int(user_input)

    elif(int(user_input) == 5):
        VV.vv2()
        a = int(user_input)
        
    elif(int(user_input) == 6):
        VV.vikas()
        a = int(user_input)

    elif(int(user_input) == 7):
        SW.anoopFrancis()
        a = int(user_input) 

    elif(int(user_input) == 8):
        SW.sai()
        a = int(user_input)

    elif(int(user_input) == 9):
        SW.neeladri()
        a = int(user_input)

    elif(int(user_input) == 10):
        SW.chandrappa()
        a = int(user_input)

    elif(int(user_input) == 11):
        SW.sowmya()
        a = int(user_input)
        
    elif(int(user_input) == 12):
        SW.haripriya()
        a = int(user_input)
        
    elif(int(user_input) == 13):
        SW.sw1()
        a = int(user_input)

    elif(int(user_input) == 14):
        SW.sw2()
        a = int(user_input)

    elif(int(user_input) == 15):
        SW.aravinda()
        a = int(user_input)
        
    elif(int(user_input) == 16):
        SW.pankaj()
        a = int(user_input)
        
    elif(int(user_input) == 17):
        safety.swami()
        a = int(user_input)
        
    elif(int(user_input) == 18):
        opera.all_users()
        a = int(user_input)
        
    elif(int(user_input) == 19):
        break

    else:
        print("Invalid input")

workbook.save(filename = weekly_report)

#Close workbook
workbook.close()
